"""
数据聚合入口。
collect(project_root: Path) -> dict
返回符合 §3 全部字段的 snapshot dict，所有 IO 失败均 graceful 降级。
"""

import re
from datetime import datetime, timedelta
from pathlib import Path

from .claude_md_parser import parse_claude_md
from .deliverable_scanner import scan_deliverables, scan_phase_files
from .signals_reader import read_signals


def _scan_risks(project_root: Path) -> list[dict]:
    """扫描风险跟踪表 xlsx，提取活跃风险。失败则返回空列表。"""
    risk_dir = project_root / "00_项目管理" / "02_风险与问题"
    if not risk_dir.exists():
        return []

    risks = []
    for f in risk_dir.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        try:
            from openpyxl import load_workbook
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            if ws is None:
                continue

            # 探测表头
            headers = []
            header_idx = 1
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if any(str(cell).strip() in ("风险编号", "风险描述", "严重程度", "状态", "责任人")
                       for cell in row if cell is not None):
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                    header_idx = row[0].row if hasattr(row[0], "row") else 1
                    break

            if not headers:
                continue

            id_col = next((i for i, h in enumerate(headers) if h in ("风险编号", "编号")), None)
            desc_col = next((i for i, h in enumerate(headers) if h in ("风险描述", "描述", "风险名称")), None)
            sev_col = next((i for i, h in enumerate(headers) if h in ("严重程度", "严重级别")), None)
            status_col = next((i for i, h in enumerate(headers) if h in ("状态", "风险状态")), None)
            owner_col = next((i for i, h in enumerate(headers) if h in ("责任人", "负责人")), None)

            for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
                if not row:
                    continue
                risk_id = str(row[id_col]).strip() if id_col is not None and id_col < len(row) and row[id_col] else ""
                if not risk_id:
                    continue
                status = str(row[status_col]).strip() if status_col is not None and status_col < len(row) and row[status_col] else ""
                if "关闭" in status or "已解决" in status or "完成" in status:
                    continue
                risks.append({
                    "risk_id": risk_id,
                    "title": str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else "",
                    "severity": str(row[sev_col]).strip() if sev_col is not None and sev_col < len(row) and row[sev_col] else "",
                    "status": status,
                    "owner": str(row[owner_col]).strip() if owner_col is not None and owner_col < len(row) and row[owner_col] else "",
                    "source_file": str(f.name),
                })
        except Exception:
            continue
    return risks


def _scan_changes(project_root: Path) -> list[dict]:
    """扫描变更申请 docx 文件。只提取文件名信息。"""
    changes = []
    # 00_项目管理/
    mgmt_dir = project_root / "00_项目管理"
    if mgmt_dir.exists():
        for f in mgmt_dir.rglob("*.docx"):
            if "变更" in f.name or "change" in f.name.lower():
                changes.append({"pcr_id": Path(f).stem, "file": str(f.relative_to(project_root))})
    # 各阶段/变更/
    for phase_dir in project_root.iterdir():
        if phase_dir.is_dir() and phase_dir.name.startswith(("0", "1", "2", "3", "4", "5", "6", "7")):
            change_sub = phase_dir / "变更"
            if change_sub.exists():
                for f in change_sub.rglob("*.docx"):
                    changes.append({"pcr_id": Path(f).stem, "file": str(f.relative_to(project_root))})
    return changes


def _scan_weekly_reports(project_root: Path) -> dict:
    """扫描周报 md 文件，统计数量和最后日期。"""
    report_dir = project_root / "00_项目管理" / "01_计划与进度"
    if not report_dir.exists():
        return {"count": 0, "last_date": None, "on_time_rate": None}

    reports = []
    for f in report_dir.glob("*周报*.md"):
        reports.append(str(f))

    dates = []
    for r in reports:
        m = re.search(r'(\d{4})[-_]?(\d{2})[-_]?(\d{2})', Path(r).name)
        if m:
            dates.append(f"{m.group(1)}-{m.group(2)}-{m.group(3)}")

    last_date = max(dates) if dates else None
    count = len(reports)
    # 简化：有文件则假设准时率 100%，否则 None
    on_time_rate = 1.0 if count > 0 else None
    return {"count": count, "last_date": last_date, "on_time_rate": on_time_rate}


def _read_worklog(project_root: Path) -> list[dict]:
    """读取 worklog.md 最近 20 条。"""
    path = project_root / "worklog.md"
    if not path.exists():
        return []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
        logs = []
        for line in lines:
            line = line.strip()
            if line.startswith("- ") or line.startswith("* "):
                logs.append({"summary": line[2:].strip()})
            elif line and not line.startswith("#"):
                logs.append({"summary": line})
        return logs[-20:]
    except Exception:
        return []


def _build_stakeholders(parsed: dict, signals: dict) -> list[dict]:
    """合并 CLAUDE.md PM 信息与 signals meeting.held 参与者。"""
    stakeholders = []
    seen = set()

    pm_name = parsed.get("pm_name")
    if pm_name:
        stakeholders.append({"name": pm_name, "role": "金蝶PM", "last_contact": None})
        seen.add(pm_name)

    client_pm = parsed.get("client_pm")
    if client_pm and client_pm not in seen:
        stakeholders.append({"name": client_pm, "role": "客户PM", "last_contact": None})
        seen.add(client_pm)

    for m in signals.get("by_type", {}).get("meeting.held", []):
        payload = m.get("payload", {})
        for p in payload.get("participants", []):
            if not p:
                continue
            ts = m.get("ts")
            if p not in seen:
                stakeholders.append({"name": p, "role": "干系人", "last_contact": ts})
                seen.add(p)
            else:
                for s in stakeholders:
                    if s["name"] == p:
                        # 保留更晚的日期
                        if ts and (s["last_contact"] is None or ts > s["last_contact"]):
                            s["last_contact"] = ts

    return stakeholders


def _meetings_this_week(signals: dict) -> list[dict]:
    """从 signals 提取本周 meeting.held events（周一至今日）。"""
    now = datetime.now().astimezone()
    week_start = now - timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)

    meetings = []
    for m in signals.get("by_type", {}).get("meeting.held", []):
        ts_str = m.get("ts", "")
        try:
            ts = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
            if ts.tzinfo is None:
                ts = ts.astimezone()
            if ts >= week_start:
                meetings.append({
                    "meeting_id": m.get("id", ""),
                    "ts": ts_str,
                    "summary": m.get("payload", {}).get("summary", ""),
                    "participants": m.get("payload", {}).get("participants", []),
                })
        except Exception:
            continue

    meetings.sort(key=lambda x: x["ts"], reverse=True)
    return meetings[:5]


def _apply_signals_to_deliverables(deliverables: dict, signals: dict) -> dict:
    """
    根据 signals.jsonl 中的 deliverable.completed 事件
    更新 cells 的 signed_off 和 status。
    """
    cells = deliverables.get("cells", [])
    stats = dict(deliverables.get("stats", {}))

    completed_events = signals.get("by_type", {}).get("deliverable.completed", [])
    for event in completed_events:
        payload = event.get("payload", {})
        phase_id = payload.get("phase_id", "")
        name = payload.get("name", "")
        signed_off = payload.get("signed_off", False)

        for cell in cells:
            cell_phase = cell.get("phase", "")
            cell_name = cell.get("name", "")
            # 去掉扩展名进行模糊匹配
            cell_name_stem = Path(cell_name).stem if "." in cell_name else cell_name
            name_stem = Path(name).stem if "." in name else name

            phase_match = (
                phase_id == cell_phase
                or phase_id in cell_phase
                or cell_phase in phase_id
            )
            name_match = (
                cell_name == name
                or cell_name_stem == name_stem
                or name in cell_name
                or cell_name_stem in name
            )
            if phase_match and name_match:
                cell["signed_off"] = signed_off
                old_status = cell.get("status", "pending")
                if signed_off and old_status != "done":
                    cell["status"] = "done"
                    stats["done"] = stats.get("done", 0) + 1
                    if old_status in stats and stats[old_status] > 0:
                        stats[old_status] -= 1
                break

    return {"cells": cells, "stats": stats}


def _determine_current_phase(phases: list[dict]) -> str:
    """从 phases 列表推断当前阶段名称。"""
    for p in phases:
        status = p.get("status", "").lower()
        if "进行" in status or "in_progress" in status:
            return p.get("name", "")
    #  fallback：找最后一个已完成的下一个，或最后一个
    for i, p in enumerate(phases):
        status = p.get("status", "").lower()
        if "未开始" in status or "待开始" in status:
            if i > 0:
                return phases[i - 1].get("name", "")
            return p.get("name", "")
    if phases:
        return phases[-1].get("name", "")
    return ""


def _day_count(start_date_str: str) -> int:
    """计算从 start_date 到今日的天数。"""
    if not start_date_str:
        return 0
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            start = datetime.strptime(start_date_str.strip(), fmt)
            return (datetime.now() - start).days
        except Exception:
            continue
    return 0


def collect(project_root: Path) -> dict:
    """
    聚合项目多源数据，返回 snapshot dict。
    所有字段均存在（可能为空/None），保证下游消费无需 KeyError 处理。
    所有 IO 失败均内部消化，永不抛异常到顶层。
    """
    snapshot = {
        "project_name": "",
        "pm_name": "",
        "client_pm": "",
        "product_type": "",
        "contract_amount": "",
        "contract_days": None,
        "start_date": "",
        "plan_go_live": "",
        "current_phase": "",
        "day_count": 0,
        "phases": [],
        "milestones": [],
        "activity_log": [],
        "deliverables": {"cells": [], "stats": {"done": 0, "pending": 0, "missing": 0, "optional": 0}},
        "signals": {"events": [], "by_baseline": {}, "by_type": {}, "count": 0},
        "risks": [],
        "changes": [],
        "weekly_reports": {"count": 0, "last_date": None, "on_time_rate": None},
        "worklog": [],
        "stakeholders": [],
        "meetings_this_week": [],
        "meta": {
            "generated_at": datetime.now().isoformat(),
            "version": "3.0.0",
            "errors": [],
        },
    }

    # ── 1. CLAUDE.md ──
    claude_md_path = project_root / "CLAUDE.md"
    parsed = {}
    if claude_md_path.exists():
        try:
            parsed = parse_claude_md(claude_md_path)
            snapshot["project_name"] = parsed.get("project_name") or ""
            snapshot["pm_name"] = parsed.get("pm_name") or ""
            snapshot["client_pm"] = parsed.get("client_pm") or ""
            snapshot["product_type"] = parsed.get("product_type") or ""
            snapshot["contract_amount"] = parsed.get("contract_amount") or ""
            snapshot["contract_days"] = parsed.get("contract_days")
            snapshot["start_date"] = parsed.get("start_date") or ""
            snapshot["plan_go_live"] = parsed.get("plan_go_live") or ""
            snapshot["phases"] = parsed.get("phases") or []
            snapshot["milestones"] = parsed.get("milestones") or []
            snapshot["activity_log"] = parsed.get("activity_log") or []
            snapshot["current_phase"] = _determine_current_phase(snapshot["phases"])
            snapshot["day_count"] = _day_count(snapshot["start_date"])
        except Exception as e:
            snapshot["meta"]["errors"].append(f"CLAUDE.md parse error: {e}")

    # ── 2. Signals ──
    try:
        signals = read_signals(project_root)
        snapshot["signals"] = signals
    except Exception as e:
        snapshot["meta"]["errors"].append(f"signals read error: {e}")
        signals = snapshot["signals"]

    # ── 3. Deliverables ──
    try:
        deliverables = scan_deliverables(project_root)
        deliverables = _apply_signals_to_deliverables(deliverables, signals)
        snapshot["deliverables"] = deliverables
    except Exception as e:
        snapshot["meta"]["errors"].append(f"deliverables scan error: {e}")

    # ── 4. Risks ──
    try:
        snapshot["risks"] = _scan_risks(project_root)
    except Exception as e:
        snapshot["meta"]["errors"].append(f"risks scan error: {e}")

    # ── 5. Changes ──
    try:
        snapshot["changes"] = _scan_changes(project_root)
    except Exception as e:
        snapshot["meta"]["errors"].append(f"changes scan error: {e}")

    # ── 6. Weekly reports ──
    try:
        snapshot["weekly_reports"] = _scan_weekly_reports(project_root)
    except Exception as e:
        snapshot["meta"]["errors"].append(f"weekly reports scan error: {e}")

    # ── 7. Worklog ──
    try:
        snapshot["worklog"] = _read_worklog(project_root)
    except Exception as e:
        snapshot["meta"]["errors"].append(f"worklog read error: {e}")

    # ── 8. Stakeholders & meetings ──
    try:
        snapshot["stakeholders"] = _build_stakeholders(parsed, signals)
        snapshot["meetings_this_week"] = _meetings_this_week(signals)
    except Exception as e:
        snapshot["meta"]["errors"].append(f"stakeholders/meetings error: {e}")

    return snapshot
