"""
扫描 7 阶段目录 + 读取必选交付物清单，生成 DeliverableCell 列表。
§7 定义了交付物管理网格的数据结构。
"""

import re
from pathlib import Path


# 7 个业务阶段目录 → 内部简称映射
_PHASE_DIR_MAP = {
    "01_启动阶段": "启动",
    "02_需求阶段": "需求",
    "03_方案阶段": "方案",
    "04_构建阶段": "构建",
    "05_测试阶段": "测试",
    "06_上线阶段": "上线",
    "07_验收阶段": "验收",
}

_EXCLUDED_PATTERNS = (
    ".DS_Store",
    "Thumbs.db",
    ".gitignore",
    "~$",
)


def _is_excluded(filename: str) -> bool:
    """排除系统文件、临时文件。"""
    if filename.startswith(".") and filename != ".gitignore":
        return True
    for pat in _EXCLUDED_PATTERNS:
        if pat in filename:
            return True
    return False


def _read_mandatory_xlsx(path: Path) -> list[dict]:
    """用 openpyxl 读取必选清单 xlsx，返回 [{'phase': str, 'name': str, 'required': bool}, ...]"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    items = []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        # 尝试定位表头行
        header_row_idx = 1
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(str(cell).strip() in ("阶段", "phase", "Phase", "交付物名称", "名称", "name", "Name")
                   for cell in row if cell is not None):
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                header_row_idx = row[0].row if hasattr(row[0], "row") else 1
                break

        if not headers:
            # 按位置推断：第一列=阶段，第二列=名称
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue
                phase = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if row[1] else ""
                if name:
                    items.append({"phase": phase, "name": name, "required": True})
            return items

        # 定位列索引
        phase_col = next((i for i, h in enumerate(headers) if h in ("阶段", "phase", "Phase")), None)
        name_col = next((i for i, h in enumerate(headers) if h in ("交付物名称", "名称", "交付物", "name", "Name")), None)
        req_col = next((i for i, h in enumerate(headers) if h in ("是否必选", "必选", "required", "Required")), None)

        if name_col is None:
            name_col = 1 if len(headers) > 1 else 0
        if phase_col is None:
            phase_col = 0

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or len(row) <= max(phase_col, name_col):
                continue
            phase = str(row[phase_col]).strip() if phase_col is not None and row[phase_col] is not None else ""
            name = str(row[name_col]).strip() if name_col is not None and row[name_col] is not None else ""
            if not name:
                continue
            required = True
            if req_col is not None and req_col < len(row) and row[req_col] is not None:
                required = str(row[req_col]).strip().lower() in ("是", "true", "yes", "1", "必选", "required")
            items.append({"phase": phase, "name": name, "required": required})
    except Exception:
        return []

    return items


def _read_mandatory_md(path: Path) -> list[dict]:
    """读取 markdown 格式的必选清单。"""
    try:
        from .claude_md_parser import _parse_markdown_table
        rows = _parse_markdown_table(path.read_text(encoding="utf-8").splitlines())
    except Exception:
        return []

    items = []
    for row in rows:
        phase = row.get("阶段", row.get("phase", "")).strip()
        name = row.get("交付物名称", row.get("名称", row.get("name", ""))).strip()
        if name:
            items.append({"phase": phase, "name": name, "required": True})
    return items


def read_mandatory_list(project_root: Path) -> list[dict]:
    """
    按优先级读取必选交付物清单：
    1. 00_项目管理/简化版必选交付物清单.xlsx
    2. 00_项目管理/必选交付物清单.md
    3. 返回空列表（由调用方自行 fallback）
    """
    xlsx_path = project_root / "00_项目管理" / "简化版必选交付物清单.xlsx"
    if xlsx_path.exists():
        items = _read_mandatory_xlsx(xlsx_path)
        if items:
            return items

    md_path = project_root / "00_项目管理" / "必选交付物清单.md"
    if md_path.exists():
        items = _read_mandatory_md(md_path)
        if items:
            return items

    return []


def scan_phase_files(project_root: Path) -> dict:
    """
    扫描 7 个阶段目录，返回 {phase_name: [relative_path, ...]}。
    目录不存在时对应值为空列表，永不抛异常。
    """
    result: dict = {}
    for dir_name, phase_name in _PHASE_DIR_MAP.items():
        phase_dir = project_root / dir_name
        files = []
        if phase_dir.exists() and phase_dir.is_dir():
            try:
                for f in phase_dir.rglob("*"):
                    if f.is_file() and not _is_excluded(f.name):
                        try:
                            files.append(str(f.relative_to(project_root)))
                        except ValueError:
                            files.append(str(f))
            except Exception:
                pass
        result[phase_name] = files
    return result


def scan_deliverables(project_root: Path) -> dict:
    """
    综合扫描阶段目录 + 必选清单，生成交付物网格数据。
    返回：
        {
            "cells": [DeliverableCell, ...],
            "stats": {"done": int, "pending": int, "missing": int, "optional": int}
        }
    所有 IO 失败均降级为只返回已有文件或空列表。
    """
    cells = []
    stats = {"done": 0, "pending": 0, "missing": 0, "optional": 0}

    try:
        existing = scan_phase_files(project_root)
    except Exception:
        existing = {name: [] for name in _PHASE_DIR_MAP.values()}

    try:
        mandatory = read_mandatory_list(project_root)
    except Exception:
        mandatory = []

    if mandatory:
        # 以清单为主，生成完整网格
        for item in mandatory:
            phase = item.get("phase", "")
            name = item.get("name", "")
            required = item.get("required", True)
            if not name:
                continue

            # 模糊匹配现有文件
            file_path = None
            for fp in existing.get(phase, []):
                # 简单匹配：文件名包含交付物名称，或交付物名称包含文件名（去除扩展名）
                fp_name = Path(fp).stem
                if name in fp or fp_name in name or name in fp_name:
                    file_path = fp
                    break

            if file_path:
                status = "pending"  # aggregator 可能根据 signals 更新为 done
            elif required:
                status = "missing"
            else:
                status = "optional"

            cells.append({
                "phase": phase,
                "name": name,
                "required": required,
                "status": status,
                "file_path": file_path,
                "signed_off": False,
            })
            stats[status] = stats.get(status, 0) + 1

        # 清单未覆盖的已有文件标记为 optional
        covered = {(c["phase"], c["name"]) for c in cells}
        for phase, files in existing.items():
            for fp in files:
                name = Path(fp).name
                if (phase, name) not in covered:
                    cells.append({
                        "phase": phase,
                        "name": name,
                        "required": False,
                        "status": "optional",
                        "file_path": fp,
                        "signed_off": False,
                    })
                    stats["optional"] = stats.get("optional", 0) + 1
    else:
        # 无清单时，所有已有文件视为 pending（或 optional）
        for phase, files in existing.items():
            for fp in files:
                cells.append({
                    "phase": phase,
                    "name": Path(fp).name,
                    "required": False,
                    "status": "pending",
                    "file_path": fp,
                    "signed_off": False,
                })
                stats["pending"] = stats.get("pending", 0) + 1

    return {"cells": cells, "stats": stats}
